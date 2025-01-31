import io
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse

def get_competitor_name(url: str, used_domains: Dict[str, int]) -> str:
    """Generate a unique competitor name based on domain."""
    parsed_url = urlparse(url)
    domain = parsed_url.netloc.replace('www.', '')
    if domain not in used_domains:
        used_domains[domain] = 0
    used_domains[domain] += 1
    if used_domains[domain] > 1:
        return f"comp_{domain}_{used_domains[domain]}"
    return f"comp_{domain}"

def validate_urls(urls: List[str]) -> bool:
    """Validate list of URLs."""
    import re
    url_pattern = re.compile(
        r'^https?://'
        r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+[A-Z]{2,6}\.?|'
        r'localhost|'
        r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})'
        r'(?::\d+)?'
        r'(?:/?|[/?]\S+)$', re.IGNORECASE)

    return all(url_pattern.match(url) for url in urls if url)

def create_excel_report(analysis_results: Dict[str, Any]) -> bytes:
    """Create Excel report from analysis results without pandas."""
    output = io.BytesIO()
    workbook = Workbook()

    # --- Recommendation Overview Sheet ---
    overview_sheet = workbook.active
    overview_sheet.title = "Recommendation Overview"
    overview_headers = ["Entity", "Relevance Score", "Reasoning"]
    overview_sheet.append(overview_headers)

    overview_data = []
    recommendation_overview = analysis_results.get("recommendation_overview")
    if recommendation_overview:
        for entity_selection in recommendation_overview.selected_entities:
            overview_data.append([
                entity_selection.entity_name,
                entity_selection.relevance_score,
                entity_selection.reasoning
            ])
    for row in overview_data:
        overview_sheet.append(row)
    adjust_column_width(overview_sheet, overview_headers)

    # --- Entity Analysis Sheet ---
    entity_sheet = workbook.create_sheet("Entity Analysis")
    entity_headers = ["Source", "Entity", "Type", "Salience", "Sentiment Score", "Sentiment Magnitude", "Mentions"]
    entity_sheet.append(entity_headers)

    entity_data = []
    for source, analysis in [("Client Page - " + analysis_results.get("client_url", ""), analysis_results["client_analysis"])] + [
        (f"Competitor - {url}", comp) 
        for url, comp in zip(analysis_results.get("competitor_names", {}).keys(), analysis_results["competitive_analyses"])
    ]:
        for entity_name, data in analysis["entities"].items():
            entity_data.append([
                source,
                entity_name,
                data["type"],
                data["salience"],
                data["sentiment"]["score"],
                data["sentiment"]["magnitude"],
                ", ".join(data["mentions"])
            ])
    for row in entity_data:
        entity_sheet.append(row)
    adjust_column_width(entity_sheet, entity_headers)

    # --- Keyword Analysis Sheet ---
    keyword_sheet = workbook.create_sheet("Keyword Analysis")
    keyword_headers = ["Source", "Keyword", "Density", "Count", "TF-IDF", "Phrase Count"]
    keyword_sheet.append(keyword_headers)

    keyword_data = []
    for source, analysis in [("Client Page - " + analysis_results.get("client_url", ""), analysis_results["client_analysis"])] + [
        (f"Competitor - {url}", comp) 
        for url, comp in zip(analysis_results.get("competitor_names", {}).keys(), analysis_results["competitive_analyses"])
    ]:
        for keyword, data in analysis["keyword_analysis"].items():
            keyword_data.append([
                source,
                keyword,
                data["density"],
                data["count"],
                data["tf_idf"],
                data["phrase_counts"]
            ])
    for row in keyword_data:
        keyword_sheet.append(row)
    adjust_column_width(keyword_sheet, keyword_headers)

    # --- Missing Entities Sheet ---
    missing_entities_sheet = workbook.create_sheet("Missing Entities")
    missing_entities_headers = ["Entity", "Type", "Relevance", "Reasoning"]
    missing_entities_sheet.append(missing_entities_headers)

    missing_entities = []
    for entity_name, data in analysis_results["comparison_results"]["missing_entities"].items():
        missing_entities.append([
            entity_name,
            data["type"],
            data["relevance"],
            data["reasoning"]
        ])
    for row in missing_entities:
        missing_entities_sheet.append(row)
    adjust_column_width(missing_entities_sheet, missing_entities_headers)

    # --- Missing Keywords Sheet ---
    missing_keywords_sheet = workbook.create_sheet("Missing Keywords")
    missing_keywords_headers = ["Keyword", "Competitor URL", "Density", "Count", "TF-IDF"]
    missing_keywords_sheet.append(missing_keywords_headers)

    missing_keywords = []
    for keyword, data in analysis_results["comparison_results"]["missing_keywords"].items():
        for url, metrics in data["competitors"].items():
            missing_keywords.append([
                keyword,
                url,
                metrics["density"],
                metrics["count"],
                metrics["tf_idf"]
            ])
    for row in missing_keywords:
        missing_keywords_sheet.append(row)
    adjust_column_width(missing_keywords_sheet, missing_keywords_headers)

    # --- Document Sentiment Sheet ---
    sentiment_sheet = workbook.create_sheet("Document Sentiment")
    sentiment_headers = ["Source", "Score", "Magnitude"]
    sentiment_sheet.append(sentiment_headers)

    sentiment_data = [
        [
            "Client Page - " + analysis_results.get("client_url", ""),
            analysis_results["client_analysis"]["document_sentiment"]["score"],
            analysis_results["client_analysis"]["document_sentiment"]["magnitude"]
        ]
    ]

    for url, comp in zip(analysis_results.get("competitor_names", {}).keys(), analysis_results["competitive_analyses"]):
        sentiment_data.append([
            f"Competitor - {url}",
            comp["document_sentiment"]["score"],
            comp["document_sentiment"]["magnitude"]
        ])
    for row in sentiment_data:
        sentiment_sheet.append(row)
    adjust_column_width(sentiment_sheet, sentiment_headers)

    # --- AI Recommendations Sheet ---
    recommendations_sheet = workbook.create_sheet("AI Recommendations")
    recommendations_headers = ["Entity", "Recommendation"]
    recommendations_sheet.append(recommendations_headers)

    recommendation_data = []
    for entity, recommendation in analysis_results.get("entity_recommendations", {}).items():
        recommendation_data.append([
            entity,
            recommendation.to_markdown if hasattr(recommendation, 'to_markdown') else recommendation
        ])
    for row in recommendation_data:
        recommendations_sheet.append(row)
    adjust_column_width(recommendations_sheet, recommendations_headers)

    workbook.save(output)
    output.seek(0)
    return output.getvalue()

def adjust_column_width(sheet, headers):
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = max(len(str(header)), max(len(str(cell.value)) if cell.value else 0 for cell in sheet[column_letter]))
        adjusted_width = (max_length + 2)
        adjusted_width = 75 if adjusted_width > 75 else adjusted_width # if adjusted_width > 75, make it 75 to prevent column width from being too large
        sheet.column_dimensions[column_letter].width = adjusted_width